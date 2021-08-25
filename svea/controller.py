import codecs
import shutil
import time
from pathlib import Path
import os
import sys
import openpyxl

import subprocess
import webbrowser

from ctdpy.core import session as ctdpy_session
from ctdpy.core.utils import generate_filepaths, get_reversed_dictionary
from sharkpylib.qc.qc_default import QCBlueprint

from bokeh.plotting import curdoc
from ctdvis  import session as ctdvis_session
from ctd_processing.former_processing import CtdProcessing

from svea import exceptions

import logging
import logging.config
import logging.handlers

SHARK_PACKAGES = ['sharkpylib', 'ctdpy', 'ctdvis']


class SveaSteps:
    def __init__(self):
        self.sbe_processing = False
        self.create_metadata_file = False
        self.create_standard_format = False
        self.perform_automatic_qc = False
        self.open_visual_qc = False
        self.send_files_to_ftp = False
        self.import_to_lims = False
        self.create_station_plots = False


class CommonFiles:
    def change_location(self, directory):
        """
        Files will be copied to new location. Option to overwrite.
        :param directory:
        :param overwrite:
        :return:
        """
        if self._file_paths is None:
            raise exceptions.PathError(f'Missing paths for {self._title}')
        print('directory', directory)
        directory = Path(directory)
        if '.' in directory.name:
            text = f'Path is not a directory: {directory}'
            self.logger.error(text)
            raise exceptions.PathError(text)
        if not directory.exists():
            os.makedirs(directory)

        self.logger.info(f'Copying files. Permission to overwrite is set to {self.allow_overwrite}')
        file_paths = []
        for file_path in self._file_paths:
            file_name = file_path.name
            new_file_path = Path(directory, file_name)
            file_paths.append(new_file_path)
            if not self.allow_overwrite and new_file_path.exists():
                continue
            if file_path == new_file_path:
                continue
            shutil.copyfile(file_path, new_file_path)
        self._file_paths = file_paths


class SveaController:
    def __init__(self, logger=None):
        self.logger = get_logger(logger)
        self.logger.debug('In SveaController')
        
        self.dirs = {}

        self.dirs['working'] = None
        self.dirs['cnv_files'] = None
        self.dirs['raw_files'] = None
        self.dirs['standard_files'] = None
        self.dirs['standard_files_qc'] = None

        self.bokeh_visualize_setting = 'smhi_vis'
        self.bokeh_visualize_setting = 'deep_vis'

        self.bokeh_server_venv_path = Path(Path(__file__).parent.parent, 'venv')
        self.bokeh_server_directory = Path(Path(__file__).parent.parent, 'bokeh_server')

        self._steps = SveaSteps()

        self._ctd_processing_object = CtdProcessing(logger=self.logger)

        self._raw_files_object = RawFiles(logger=logger)

        self._metadata_object = Metadata(logger=self.logger)

        self._sensorinfo_object = SensorInfo(logger=self.logger)

        self._metadata_file_object = MetadataFile(logger=self.logger)
        self._metadata_file_object.metadata_object = self._metadata_object
        self._metadata_file_object.sensor_info_object = self._sensorinfo_object

        self._cnv_files_object = CNVfiles(logger=self.logger)
        
        self._standard_files_object = ProfileStandardFormatFiles(logger=self.logger)

        self._create_metadata_file_object = CreateMetadataFile(logger=self.logger)
        self._create_metadata_file_object.metadata_file_object = self._metadata_file_object
        self._create_metadata_file_object.cnv_files_object = self._cnv_files_object

        self._create_standard_files_object = CreateStandardFormatFiles(logger=self.logger)
        self._create_standard_files_object.metadata_file_object = self._metadata_file_object
        self._create_standard_files_object.cnv_files_object = self._cnv_files_object
        
        self._automatic_qc_object = AutomaticQC(logger=self.logger)
        self._automatic_qc_object.standard_files_object = self._standard_files_object

        self._visual_qc_object = VisualQC(logger=self.logger)

        self.logger.info('SveaController instance created!')
        
    def __repr__(self):
        dirs_list = []
        for key, value in self.dirs.items():
            dirs_list.append(f'{key}: {value}')
        return '\n'.join(dirs_list)

    def _assert_directory(self):
        if not self.dirs['working']:
            text = 'Working directory is not set'
            self.logger.error(text)
            raise exceptions.MissingFiles(text)
        elif not self.dirs['working'].exists():
            os.makedirs(self.dirs['working'])
            self.logger.info(f'Woring directory created: {self.dirs["working"]}')

    @property
    def working_directory(self):
        return self.dirs['working']

    @working_directory.setter
    def working_directory(self, directory):
        print('working directory', directory)
        if directory is None:
            self.dirs['working'] = None
            self.dirs['cnv_files'] = None
            self.dirs['raw_files'] = None
            self.dirs['standard_files'] = None
            self.dirs['standard_files_qc'] = None
        else:
            self.dirs['working'] = Path(directory)
            self.dirs['raw_files'] = Path(self.dirs['working'], 'raw_files')
            self.dirs['cnv_files'] = Path(self.dirs['working'], 'cnv')
            self.dirs['standard_files'] = Path(self.dirs['working'], 'standard_format')
            self.dirs['standard_files_qc'] = Path(self.dirs['working'], 'standard_format_auto_qc')

        self._metadata_file_object.file_path = self.dirs['cnv_files']
        self._create_standard_files_object.directory = self.dirs['standard_files']
        self._standard_files_object.file_paths = self.dirs['standard_files']

        self.logger.info(f'Working directory set to: {directory}')

    def set_path_working_directory(self, directory):
        self.working_directory = directory

    @property
    def metadata_file_path(self):
        return self._metadata_file_object.file_path

    @property
    def ctd_processing_options(self):
        return self._ctd_processing_object.options

    def sbe_processing(self, file_path, **kwargs):
        """
        kwargs are options that you can get from self.ctd_processing_options
        :param kwargs:
        :return:
        """
        for key, value in kwargs.items():
            setattr(self._ctd_processing_object, key, value)
        print('FILE PATH', file_path)
        self._ctd_processing_object.load_seabird_files(file_path)
        self._ctd_processing_object.run_process()

        self._assert_directory() 
        # if not self._raw_files_object.file_paths:
        #     raise exceptions.PathError('No raw files selected')
        # self._raw_files_object.change_location(self.dirs['raw_files'])
        self._steps.sbe_processing = True
        # return self.dirs['raw_files']

    def create_metadata_file(self):
        self._assert_directory()
        self._create_metadata_file_object.create_file()
        self._cnv_files_object.change_location(self.dirs['cnv_files'])
        self._steps.create_metadata_file = True
        return self.dirs['cnv_files']

    def create_standard_format(self):
        self._assert_directory()
        self._create_standard_files_object.create_files()
        self._steps.create_standard_format = True
        return self._create_standard_files_object.directory

    def perform_automatic_qc(self):
        self._assert_directory()
        self._automatic_qc_object.run_qc(self.dirs['standard_files_qc'])
        self._steps.perform_automatic_qc = True
        return self.dirs['standard_files_qc']

    def open_visual_qc(self, server_file_directory=None, venv_path=None, shark_package_root=None, **filters): 
        
        if server_file_directory:
            path = Path(server_file_directory)
            if path.exists():
                self.bokeh_server_directory = path

        if venv_path:
            self.bokeh_server_venv_path = Path(venv_path)
        if not self.bokeh_server_venv_path.exists():
            raise exceptions.PathError(f'Virtual environment not found at {self.bokeh_server_venv_path}')

        self._create_bokeh_server_source_directory(shark_package_root=shark_package_root)

        if not self.dirs['standard_files_qc']:
            raise exceptions.PathError('Path to qc standard files not set')
        if not os.listdir(self.dirs['standard_files_qc']):
            raise exceptions.MissingFiles('Missing files to visualize')

        self._visual_qc_object.set_options(data_directory=self.dirs['standard_files_qc'],
                                           visualize_setting=self.bokeh_visualize_setting,
                                           server_file_directory=self.bokeh_server_directory,
                                           venv_path=self.bokeh_server_venv_path,
                                           **filters)
        self._visual_qc_object.run()
        self._steps.open_visual_qc = True
        
    def close_visual_qc(self):
        self._visual_qc_object.kill_server()

    def _create_bokeh_server_source_directory(self, shark_package_root=None):
        if not self.bokeh_server_directory.exists():
            os.makedirs(self.bokeh_server_directory)

        # Check if shark packages are in venv that is going to be used. 
        packages_in_venv = get_paths_to_shark_packages_in_venv(self.bokeh_server_venv_path)
        packages_in_bokeh_server_root = get_shark_packages_in_bokeh_server_root(self.bokeh_server_directory)

        shark_package_source = {}
        if shark_package_root:
            shark_package_source = get_paths_to_shark_packages_in_venv(shark_package_root)

        for package in SHARK_PACKAGES:
            if package in packages_in_bokeh_server_root:
                continue
            if packages_in_venv.get(package):
                continue 
            source_directory = shark_package_source.get(package)
            if source_directory:
                # Copy package to bokeh server directory
                target_directory = Path(self.bokeh_server_directory, source_directory.name)
                shutil.copytree(source_directory, target_directory)
                continue
            raise exceptions.MissingSharkModules(str(missing))
            

    # def _check_valid_server_directory(self, d):
    #     required = ['sharkpylib', 'ctdpy', 'ctdvis']
    #     names = os.listdir(d)
    #     missing = []
    #     for req in required:
    #         if req in sys.modules:
    #             continue
    #         if req in names: 
    #             continue
    #         missing.append(req)
    #     return missing
    #     # if missing:
    #     #     raise exceptions.MissingSharkModules(str(missing))

    def send_files_to_ftp(self):
        self._steps.send_files_to_ftp = True

    def import_to_lims(self):
        self._steps.import_to_lims = True

    def create_station_plots(self):
        self._steps.create_station_plots = True

    @property
    def metadata(self):
        return self._metadata_object.get()

    @metadata.setter
    def metadata(self, metadata):
        self._metadata_object.set(metadata)  # option to update as well (other method)

    @property
    def raw_files(self):
        return self._raw_files_object.file_paths

    @raw_files.setter
    def raw_files(self, paths):
        self._raw_files_object.file_paths = paths

    def set_path_raw_files(self, paths):
        self.raw_files = paths

    @property
    def cnv_files(self):
        return self._cnv_files_object.file_paths

    @cnv_files.setter
    def cnv_files(self, paths):
        self._cnv_files_object.file_paths = paths

    def set_path_cnv_files(self, paths):
        self.cnv_files = paths

    @property
    def standard_format_files(self):
        return self._standard_files_object.file_paths

    @standard_format_files.setter
    def standard_format_files(self, paths):
        self._standard_files_object.file_paths = paths

    def set_path_standard_format_files(self, paths):
        self.standard_format_files = paths
        
    @property
    def standard_format_files_qc(self):
        return self.dirs['standard_files_qc']
    
    @standard_format_files_qc.setter
    def standard_format_files_qc(self, path):
        try:
            path = Path(path)
            if path.exists():
                self.dirs['standard_files_qc'] = path
        except:
            pass
            
    def set_path_standard_format_files_qc(self, paths):
        self.standard_format_files_qc = paths

    def set_overwrite_permission(self, overwrite):
        if type(overwrite) != bool:
            text = 'Overwrite permission needs to be of type boolean'
            self.logger.error(text)
            raise exceptions.DtypeError(text)
        self._metadata_file_object.allow_overwrite = overwrite
        self._cnv_files_object.allow_overwrite = overwrite
        self._create_metadata_file_object.allow_overwrite = overwrite
        self._create_standard_files_object.allow_overwrite = overwrite
        self._automatic_qc_object.allow_overwrite = overwrite
        self._ctd_processing_object.overwrite = overwrite

    def reset_paths(self):
        self.raw_files = None
        self.cnv_files = None
        self.standard_format_files = None


class RawFiles(CommonFiles):
    def __init__(self, logger=None):
        self._title = 'raw files'
        self.logger = get_logger(logger)
        self._file_paths = None
        self.allow_overwrite = False

    @property
    def file_paths(self):
        return self._file_paths

    @file_paths.setter
    def file_paths(self, file_paths):
        if file_paths is None:
            self._file_paths = None
            return
        suffix_list = ['bl', 'btl', 'hdr', 'hex', 'ros', 'XMLCON', 'CON']
        print('=== file_paths', file_paths)
        if type(file_paths) in [str, Path]:
            file_paths = Path(file_paths)
            if file_paths.is_dir():
                self._file_paths = [Path(file_path) for file_path in list(generate_filepaths(file_paths,
                                                                                             pattern_list=[f'.{suffix}' for suffix in suffix_list],
                                                                                             only_from_dir=True))]
            else:
                raise exceptions.PathError('Path given to RawFiles is not a directory')
        else:
            self._file_paths = [Path(file_path) for file_path in file_paths if file_path.split('.')[-1] in suffix_list]


class Metadata:
    def __init__(self, logger=None):
        self.logger = get_logger(logger)
        self.data = {}
        self.allow_overwrite = False
        
    def add(self, metadata):
        if type(metadata) != dict:
            raise exceptions.DtypeError('metadata should be of type dict')
        self.data.update(metadata)
        
    def get(self):
        return self.data

    def set(self, metadata):
        if type(metadata) != dict:
            raise exceptions.DtypeError('metadata should be of type dict')
        self.data = metadata


class SensorInfo:
    def __init__(self, logger=None):
        self.logger = get_logger(logger)
        self.data = None

    def load_xlsx_sheet(self, file_path, sheet_name):
        wb = openpyxl.load_workbook(filename=file_path)
        if sheet_name not in wb.sheetnames:
            text = f'No worksheet named {sheet_name} in file {file_path}'
            self.logger.error(text)
            raise exceptions.PathError(text)
        ws = wb[sheet_name]
        self.data = {}
        for r, row in enumerate(ws):
            if r < 2:
                continue
            for c, cell in enumerate(row):
                if c == 0:
                    continue
                value = cell.value
                if value is None:
                    value = ''
                self.data[cell.coordinate] = str(value)

    def load_txt(self, file_path, **kwargs):
        self.data = {}
        with codecs.open(file_path, **kwargs) as fid:
            for r, line in enumerate(fid):
                split_line = line.strip('\n\r').split('\t')
                str_nr = 66
                for value in split_line:
                    col_row_str = f'{chr(str_nr)}{str(r+3)}'
                    self.data[col_row_str] = value
                    str_nr += 1


class MetadataFile:
    def __init__(self, logger=None):
        self._title = 'metadata file'
        self.logger = get_logger(logger)
        self._file_path = None
        self.metadata_object = None  # source for update
        self.sensor_info_object = None
        self.allow_overwrite = False

    @property
    def file_path(self):
        return self._file_path

    @file_path.setter
    def file_path(self, file_path):
        if file_path is None:
            self._file_path = None
            return
        path = Path(file_path)
        # file_path can be both file path and directory. if no xlsx-file is found in directory file_path is sett to actual file path when metadata file is created.
        if path.is_dir():
            for file_name in os.listdir(path):
                if file_name.endswith('.xlsx'):
                    path = Path(path, file_name)
                    break


        self._file_path = path
    
    @property
    def metadata(self):
        if not self.metadata_object:
            self.logger.info('No metadata added to update!')
            return {}
        return self.metadata_object.get()
    
    @property
    def overwrite_metadata(self):
        if not self.metadata_object:
            return False
        return self.metadata_object.allow_overwrite

    def change_location(self, directory):
        """
        Files will be copied to new location. Option to overwrite.
        :param directory:
        :param overwrite:
        :return:
        """
        self._assert_file_exists()
        directory = Path(directory)
        if not directory.is_dir():
            text = f'Path is not a directory: {directory}'
            self.logger.error(text)
            raise exceptions.PathError(text)
        if not directory.exists():
            os.makedirs(directory)

        new_file_path = Path(directory, self._file_path.name)
        if not self.allow_overwrite and new_file_path.exists():
            self.logger.warning(f'Permission to overwrite metadata file is set to {self.allow_overwrite}. File is not copied.')

        shutil.copyfile(self._file_path, new_file_path)
        self._file_path = new_file_path

    def add_sensorinfo_from_file(self, file_path, sheet_name=None):
        self._assert_file_exists()
        file_path = Path(file_path)
        if file_path.suffix == '.txt':
            self.sensor_info_object.load_txt(file_path)
        elif file_path.suffix == '.xlsx':
            self.sensor_info_object.load_xlsx_sheet(file_path, sheet_name=sheet_name)

        wb = openpyxl.load_workbook(self._file_path)
        ws = wb['Sensorinfo']
        for key, value in self.sensor_info_object.data.items():
            ws[key] = value
        wb.save(self._file_path)

    def _assert_file_exists(self):
        if not self._file_path.exists():
            text = f'Metadata file does not exist: {self._file_path}'
            self.logger.error(text)
            raise exceptions.MissingFiles(text)
            
            
class CNVfiles(CommonFiles):
    def __init__(self, logger=None):
        self._title = 'cnv files'
        self.logger = get_logger(logger)
        self._file_paths = None
        self.allow_overwrite = False

    @property
    def file_paths(self):
        return self._file_paths

    @file_paths.setter
    def file_paths(self, file_paths):
        if file_paths is None:
            self._file_paths = None
            return
        if type(file_paths) in [str, Path]:
            file_paths = Path(file_paths)
            if file_paths.is_dir():
                self._file_paths = [Path(file_path) for file_path in list(generate_filepaths(file_paths,
                                                                                        pattern_list=['.cnv'],
                                                                                        only_from_dir=True))]
            else:
                self._file_paths = [file_paths]
        else:
            self._file_paths = [Path(file_path) for file_path in file_paths if file_path.endswith('.cnv')]


class ProfileStandardFormatFiles(CommonFiles):
    def __init__(self, logger=None):
        self._title = 'standard format files'
        self.logger = get_logger(logger)
        self._file_paths = None
        self.allow_overwrite = False

    @property
    def file_paths(self):
        return self._file_paths

    @file_paths.setter
    def file_paths(self, file_paths):
        print('SETTING PATH', file_paths)
        if file_paths is None:
            self._file_paths = None
            return
        if isinstance(file_paths, str) or isinstance(file_paths, Path):
            file_paths = Path(file_paths)
            if file_paths.is_dir():
                file_paths = [Path(file_path) for file_path in list(generate_filepaths(file_paths,
                                                                                        pattern_list=['.txt'],
                                                                                        only_from_dir=True))]
            else:
                file_paths = [file_paths]
        else:
            file_paths = [Path(file_path) for file_path in file_paths if str(file_path).endswith('.txt')]
        self._file_paths = [path for path in file_paths if str(path.name).startswith('ctd_profile')]
        print('LEN _file_paths', len(self._file_paths))


class CreateMetadataFile:
    def __init__(self, logger=None):
        self.logger = get_logger(logger)

        self.metadata_file_object = None
        self.cnv_files_object = None

        self.allow_overwrite = False

        self.session = None

    def create_file(self):       
        self._assert_metadata_info_is_present()
        self._assert_cnv_files_info_is_present()
        self.session = ctdpy_session.Session(filepaths=self.cnv_files_object.file_paths,
                                             reader='smhi')

        datasets = self._get_datasets()
        dataset = datasets[0]
        self._update_metadata_in_dataset(dataset=dataset)
        self._save_file(dataset=dataset)

    def change_location(self, directory):
        self._assert_cnv_files_info_is_present()
        self._assert_metadata_info_is_present()
        self.cnv_files_object.change_location(directory, overwrite=self.allow_overwrite)
        self.metadata_file_object.change_location(directory, overwrite=self.allow_overwrite)

    def _get_datasets(self):
        start_time = time.time()
        datasets = self.session.read()
        self.logger.debug(f'{len(self.cnv_files_object.file_paths)} CNV files loaded in {time.time() - start_time} seconds.')
        return datasets

    def _update_metadata_in_dataset(self, dataset=None):
        
        self.session.update_metadata(datasets=dataset,
                                     metadata=self.metadata_file_object.metadata,
                                     overwrite=self.metadata_file_object.overwrite_metadata)
        self.logger.debug('Metadata updated in dataset')

    def _save_file(self, dataset=None):
        start_time = time.time()
        save_path = self.session.save_data(dataset,
                                           writer='metadata_template',
                                           return_data_path=True)
        self.logger.debug(f'Metadata file saved in {time.time() - start_time} seconds at location {save_path}')
        source_path = Path(save_path)
        target_path = Path(self.metadata_file_object.file_path)
        if '.' not in target_path.name:
            if not target_path.exists():
                os.makedirs(target_path)
            target_path = Path(target_path, source_path.name)
        if target_path.exists() and not self.allow_overwrite:
            text = 'Metadata file already exists and overwrite is set to False'
            self.logger.error(text)
            raise exceptions.PermissionError(text)
        shutil.copyfile(source_path, target_path)
        self.metadata_file_object.file_path = target_path  # Updated file_path if it was a directory
        self.logger.debug(f'File copied to location {target_path}')
        
    def _assert_metadata_info_is_present(self):
        text = ''
        if not self.metadata_file_object:
            text = 'No metadata file object is set by user'
        elif not self.metadata_file_object.file_path:
            text = 'File path for metadata is not set'
        # elif not self.metadata_file_object.file_path.exists():
        #     text = f'File path for metadata does not exist: {self.metadata_file_object.file_path}'
        else:
            return
        self.logger.error(text)
        raise exceptions.MissingFiles(text)

    def _assert_cnv_files_info_is_present(self):
        if not self.cnv_files_object or not self.cnv_files_object.file_paths:
            text = 'No cnv files object is set by user'
            self.logger.error(text)
            raise exceptions.MissingFiles(text)
            

class CreateStandardFormatFiles:
    def __init__(self, logger=None):
        self.logger = get_logger(logger)

        self.metadata_file_object = None
        self.cnv_files_object = None

        self.allow_overwrite = False

        self._directory = None

    @property
    def directory(self):
        return self._directory

    @directory.setter
    def directory(self, directory):
        if directory is None:
            self._directory = None
            return
        directory = Path(directory)
        if '.' in directory.name:
            text = f'Path is not a directory: {directory}'
            self.logger.error(text)
            raise exceptions.PathError(text)
        self._directory = directory

    def change_source_location(self, directory, overwrite=False):
        self._assert_metadata_and_cnv()
        self.cnv_files_object.change_location(directory, overwrite=overwrite)
        self.metadata_file_object.change_location(directory, overwrite=overwrite)

    def create_files(self):
        self._assert_metadata_and_cnv()
        self._assert_directory()
        all_file_paths = self.cnv_files_object.file_paths + [self.metadata_file_object.file_path]
        all_file_paths = [str(path) for path in all_file_paths]
        session = ctdpy_session.Session(filepaths=all_file_paths,
                                        reader='smhi')

        start_time = time.time()
        datasets = session.read()
        self.logger.debug(f'{len(self.cnv_files_object.file_paths)} CNV files and one metadata file loaded in {time.time() - start_time} seconds.')
        self.datasets = datasets
        start_time = time.time()
        data_path = session.save_data(datasets,
                                      writer='ctd_standard_template',
                                      return_data_path=True,
                                      # save_path=save_directory,
                                      )

        self.logger.warning(f'Permission to overwrite existing standard format files is set to {self.allow_overwrite}')
        for file_name in os.listdir(data_path):
            source_path = Path(data_path, file_name)
            target_path = Path(self._directory, file_name)
            if target_path.exists() and not self.allow_overwrite:
                continue
            shutil.copyfile(source_path, target_path)

        self.logger.debug(f"Datasets saved in {time.time() - start_time} sec at location: {data_path}. Files copied to: {self._directory}")

    def _assert_directory(self):
        if not self._directory:
            text = 'No directory for standard format files set'
            self.logger.error(text)
            raise exceptions.MissingFiles(text)
        elif not self._directory.exists():
            os.makedirs(self._directory)

    def _assert_metadata_and_cnv(self):
        self._assert_metadata_info_is_present()
        self._assert_cnv_files_info_is_present()

    def _assert_metadata_info_is_present(self):
        if not self.metadata_file_object:
            text = 'No metadata file object is set by user'
            self.logger.error(text)
            raise exceptions.MissingFiles(text)

    def _assert_cnv_files_info_is_present(self):
        if not self.cnv_files_object or not self.cnv_files_object.file_paths:
            text = 'No cnv files object is set by user'
            self.logger.error(text)
            raise exceptions.MissingFiles(text)


class AutomaticQC:
    def __init__(self, logger=None):
        self.logger = get_logger(logger)
        # self._file_paths = None
        self.allow_overwrite = False

        self.standard_files_object = None

    def run_qc(self, output_directory=None):
        files = self.standard_files_object.file_paths
        if not files:
            raise exceptions.MissingFiles('No standard files selected')
        session = ctdpy_session.Session(filepaths=files,
                                        reader='ctd_stdfmt')

        datasets = session.read()

        for data_key, item in datasets[0].items():
            # print(data_key)
            parameter_mapping = get_reversed_dictionary(session.settings.pmap, item['data'].keys())
            qc_run = QCBlueprint(item, parameter_mapping=parameter_mapping)
            qc_run()

        data_path = session.save_data(datasets,
                                      writer='ctd_standard_template', return_data_path=True,
                                      # save_path='C:/ctdpy_exports',
                                      )

        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
        for file_name in os.listdir(data_path):
            source_path = Path(data_path, file_name)
            target_path = Path(output_directory, file_name)
            if target_path.exists() and not self.allow_overwrite:
                continue
            shutil.copyfile(source_path, target_path)

        return output_directory


class VisualQC:
    def __init__(self, logger=None):
        self.logger = get_logger(logger)
        self.bokeh_server_file_name = 'run_bokeh_server.py'
        self.bokeh_server_file_path = Path()
        # self.run_bokeh_server_batch_file_path = Path(Path(__file__).parent, 'temp', 'run_bokeh_server.bat')
        # if not self.run_bokeh_server_batch_file_path.parent.exists():
        #     os.makedirs(self.run_bokeh_server_batch_file_path.parent)
        self.url_base = None
        self.lines = []

    def __repr__(self):
        str_list = ['Filter options are:']
        for s in ['month_list', 'ship_list', 'serno_min', 'serno_max']:
            str_list.append(s)
        return '\n'.join(str_list)

    def set_options(self, data_directory=None, visualize_setting='', server_file_directory=None, venv_path=None, **filters):
        template_source_path = Path(Path(__file__).parent, 'templates', 'bokeh_server_template.py')
        self.lines = []
        with open(template_source_path) as fid:
            for line in fid:
                if line.startswith('URL'):
                    self.url_base = line.split('=')[1].strip().strip('"').strip("'")
                elif line.startswith('DATA_DIR'):
                    line = f'DATA_DIR = r"{data_directory}"\n'
                elif filters.get('month_list') and line.startswith('MONTH_LIST'):
                    line = f'MONTH_LIST = {filters.get("month_list")}\n'
                elif filters.get('ship_list') and line.startswith('SHIP_LIST'):
                    line = f'SHIP_LIST = {filters.get("ship_list")}\n'
                elif filters.get('serno_min') and line.startswith('SERNO_MIN'):
                    line = f'SERNO_MIN = {filters.get("serno_min")}\n'
                elif filters.get('serno_max') and line.startswith('SERNO_MAX'):
                    line = f'SERNO_MAX = {filters.get("serno_max")}\n'
                elif visualize_setting and line.startswith('VISUALIZE_SETTINGS'):
                    line = f'VISUALIZE_SETTINGS = "{visualize_setting}"\n'
                self.lines.append(line)

        self._save_server_file(server_file_directory)
        self._create_batch_file(server_file_directory, venv_path)

    def _save_server_file(self, directory):
        if not self.lines:
            raise exceptions.SveaException
        self.bokeh_server_file_path = Path(directory, self.bokeh_server_file_name)
        with open(self.bokeh_server_file_path, 'w') as fid:
            fid.write(''.join(self.lines))

    def _create_batch_file(self, directory, venv_path):
        self.run_bokeh_server_batch_file_path = Path(directory, 'run_bokeh_server.bat')
        with open(self.run_bokeh_server_batch_file_path, 'w') as fid:
            fid.write(f'call {str(venv_path)}/Scripts/activate\n')
            fid.write(f'cd {str(self.bokeh_server_file_path.parent)}\n')
            fid.write(f'bokeh serve {self.bokeh_server_file_name}')
            
    def _run_server(self):
        self.bokeh_subprocess = subprocess.Popen(str(self.run_bokeh_server_batch_file_path), 
                                                 shell=False, stdout=subprocess.PIPE)
        
    def kill_server(self):
        if hasattr(self, 'bokeh_subprocess'):
            print('killing server')
            self.bokeh_subprocess.kill()
            print('It worked')

    def _open_webbrowser(self):
        url = self.url_base + self.bokeh_server_file_path.stem
        webbrowser.open(url=url)

    def run(self):
        self._run_server()
        self._open_webbrowser()


def get_logger(existing_logger=None):
    if existing_logger:
        return existing_logger
    if not os.path.exists('log'):
        os.makedirs('log')
    logging.config.fileConfig('logging.conf')
    logger = logging.getLogger('timedrotating')
    return logger

def get_directrory_path_for_string(root, string):
    for root, dirs, files in os.walk(root, topdown=False):
        # for name in files:
        #     print(os.path.join(root, name))
        for name in dirs:
            if name == string:
                return Path(root, name)
    return None

def get_paths_to_shark_packages_in_venv(venv): 
    paths = {}
    for package in SHARK_PACKAGES:
        path = get_directrory_path_for_string(venv, package) 
        if path:
            paths[path.name] = path
    return paths 

def get_shark_packages_in_bokeh_server_root(boke_server_root):
    packages = []
    for d in os.listdir(boke_server_root):
        if d in SHARK_PACKAGES:
            packages.append(d)
    return packages





if __name__ == '__main__':

    c = SveaController()

    if 1:
        c.working_directory = r'C:\mw\temp_svea/svea_repo'
        # c.cnv_files = r'C:\mw\data\cnv_files'
        # c.raw_files = r'C:\mw\data\sbe_raw_files'
        #
        # c.set_overwrite_permission(True)
        # c.sbe_processing()
        # c.create_metadata_file()
        # c.create_standard_format()

    if 0:
        directory = r'C:\mw\Profile\2019\SHARK_Profile_2019_BAS_DEEP\processed_data'
        server_directory = r'C:\mw\git\svea'
        venv_path = r'C:\mw\git\svea\venv'
        c.dirs['standard_files_qc'] = r'C:\mw\Profile\2019\SHARK_Profile_2019_BAS_DEEP\processed_data'
        c.open_visual_qc(server_file_directory=server_directory,
                         venv_path=venv_path,
                         month_list=[4, 5, 6])


    # s1 = SensorInfo()
    # s1.load_xlsx_sheet(r'C:\mw\Profile\2018\SHARK_Profile_2018_BAS_DEEP\received_data/CTD sensorinfo_org.xlsx', 'Sensorinfo')
    #
    # s2 = SensorInfo()
    # s2.load_txt(r'C:\mw\Profile\2018\SHARK_Profile_2018_BAS_DEEP\processed_data/sensorinfo.txt')



